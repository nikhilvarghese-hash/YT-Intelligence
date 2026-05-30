"""
Export: CSV, Excel, JSON, Markdown for comments, videos, creators, collections.
"""
import csv
import io
import json
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Comment, Video, Creator, CollectionItem

router = APIRouter()


def _get_comments(
    db: Session,
    creator_ids: Optional[List[int]] = None,
    collection_id: Optional[int] = None,
    query: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    min_likes: Optional[int] = None,
):
    q = (
        db.query(Comment, Video.title.label("video_title"), Video.url.label("video_url"), Creator.channel_name)
        .join(Video, Comment.video_id == Video.id)
        .join(Creator, Video.creator_id == Creator.id)
    )

    if collection_id:
        q = q.join(CollectionItem, CollectionItem.comment_id == Comment.id).filter(
            CollectionItem.collection_id == collection_id
        )
    if creator_ids:
        q = q.filter(Creator.id.in_(creator_ids))
    if query:
        q = q.filter(Comment.comment_text.ilike(f"%{query}%"))
    if date_from:
        q = q.filter(Comment.comment_date >= date_from)
    if date_to:
        q = q.filter(Comment.comment_date <= date_to)
    if min_likes is not None:
        q = q.filter(Comment.likes >= min_likes)

    return q.order_by(Comment.likes.desc()).limit(50000).all()


@router.get("/comments/csv")
def export_comments_csv(
    creator_ids: Optional[str] = Query(None),
    collection_id: Optional[int] = None,
    query: Optional[str] = None,
    min_likes: Optional[int] = None,
    db: Session = Depends(get_db),
):
    id_list = [int(x) for x in creator_ids.split(",") if x.strip()] if creator_ids else None
    rows = _get_comments(db, id_list, collection_id, query, min_likes=min_likes)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Creator", "Video Title", "Video URL", "Author", "Comment", "Likes", "Replies", "Date"])

    for r in rows:
        writer.writerow([
            r.channel_name,
            r.video_title,
            r.video_url,
            r.Comment.author_name or "",
            r.Comment.comment_text,
            r.Comment.likes,
            r.Comment.reply_count,
            r.Comment.comment_date.isoformat() if r.Comment.comment_date else "",
        ])

    output.seek(0)
    filename = f"comments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/comments/json")
def export_comments_json(
    creator_ids: Optional[str] = Query(None),
    collection_id: Optional[int] = None,
    query: Optional[str] = None,
    min_likes: Optional[int] = None,
    db: Session = Depends(get_db),
):
    id_list = [int(x) for x in creator_ids.split(",") if x.strip()] if creator_ids else None
    rows = _get_comments(db, id_list, collection_id, query, min_likes=min_likes)

    data = [
        {
            "creator": r.channel_name,
            "video_title": r.video_title,
            "video_url": r.video_url,
            "author_name": r.Comment.author_name,
            "author_channel_id": r.Comment.author_channel_id,
            "comment": r.Comment.comment_text,
            "likes": r.Comment.likes,
            "replies": r.Comment.reply_count,
            "date": r.Comment.comment_date.isoformat() if r.Comment.comment_date else None,
        }
        for r in rows
    ]

    content = json.dumps(data, ensure_ascii=False, indent=2)
    filename = f"comments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(
        iter([content]),
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/comments/markdown")
def export_comments_markdown(
    creator_ids: Optional[str] = Query(None),
    collection_id: Optional[int] = None,
    query: Optional[str] = None,
    min_likes: Optional[int] = None,
    db: Session = Depends(get_db),
):
    id_list = [int(x) for x in creator_ids.split(",") if x.strip()] if creator_ids else None
    rows = _get_comments(db, id_list, collection_id, query, min_likes=min_likes)

    lines = [
        f"# YouTube Comment Export\n",
        f"**Generated:** {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}  \n",
        f"**Total Comments:** {len(rows)}\n\n---\n",
    ]

    for r in rows:
        date_str = r.Comment.comment_date.strftime("%Y-%m-%d") if r.Comment.comment_date else "Unknown"
        lines.append(
            f"## {r.video_title}\n"
            f"**Creator:** {r.channel_name}  \n"
            f"**Author:** {r.Comment.author_name or 'Anonymous'}  \n"
            f"**Date:** {date_str} | **Likes:** {r.Comment.likes} | **Replies:** {r.Comment.reply_count}\n\n"
            f"> {r.Comment.comment_text}\n\n---\n"
        )

    content = "".join(lines)
    filename = f"comments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.md"
    return StreamingResponse(
        iter([content]),
        media_type="text/markdown",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/comments/excel")
def export_comments_excel(
    creator_ids: Optional[str] = Query(None),
    collection_id: Optional[int] = None,
    query: Optional[str] = None,
    min_likes: Optional[int] = None,
    db: Session = Depends(get_db),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Use CSV export instead.")

    id_list = [int(x) for x in creator_ids.split(",") if x.strip()] if creator_ids else None
    rows = _get_comments(db, id_list, collection_id, query, min_likes=min_likes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comments"

    headers = ["Creator", "Video Title", "Video URL", "Author", "Comment", "Likes", "Replies", "Date"]
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 80
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20

    for i, r in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=r.channel_name)
        ws.cell(row=i, column=2, value=r.video_title)
        ws.cell(row=i, column=3, value=r.video_url)
        ws.cell(row=i, column=4, value=r.Comment.author_name or "")
        cell = ws.cell(row=i, column=5, value=r.Comment.comment_text)
        cell.alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=6, value=r.Comment.likes)
        ws.cell(row=i, column=7, value=r.Comment.reply_count)
        ws.cell(row=i, column=8, value=r.Comment.comment_date.strftime("%Y-%m-%d") if r.Comment.comment_date else "")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"comments_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/creators/csv")
def export_creators_csv(db: Session = Depends(get_db)):
    from ..models import Video
    from sqlalchemy import func

    creators = db.query(Creator).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Channel Name", "Channel ID", "Subscribers", "Videos", "Channel URL", "Country", "Added"])

    for c in creators:
        writer.writerow([
            c.channel_name, c.channel_id, c.subscriber_count,
            c.video_count, c.channel_url, c.country or "",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=creators.csv"},
    )
